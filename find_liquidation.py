import openpyxl

path = "/home/ubuntu/upload/VIABILIDAD.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Viabilidad']

print("=== SEGUIMIENTO DE SALDO DEUDOR ===")
for r in range(14, 135):
    cuota_no = ws.cell(row=r, column=2).value
    cuota_val = ws.cell(row=r, column=3).value
    interes = ws.cell(row=r, column=4).value
    amort = ws.cell(row=r, column=5).value
    saldo = ws.cell(row=r, column=6).value
    
    # Si saldo es un string o espacio, o "ultimo pago"
    if saldo == "ultimo pago" or (isinstance(saldo, str) and "ultimo" in saldo.lower()):
        print(f"Fila {r}: Cuota {cuota_no} | Cuota Val: {cuota_val} | Interés: {interes} | Amortización: {amort} | Saldo: {saldo}")
        break
    if isinstance(saldo, (int, float)) and saldo <= 0:
        print(f"Fila {r}: Cuota {cuota_no} | Cuota Val: {cuota_val} | Interés: {interes} | Amortización: {amort} | Saldo: {saldo} (PAGO COMPLETADO)")
        break
    
    # Imprimir algunas cuotas clave
    if cuota_no in [0, 1, 2, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95]:
        print(f"Fila {r}: Cuota {cuota_no} | Cuota Val: {cuota_val} | Interés: {interes} | Amortización: {amort} | Saldo: {saldo}")
