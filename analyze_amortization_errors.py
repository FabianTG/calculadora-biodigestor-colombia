import openpyxl

path = "/home/ubuntu/upload/VIABILIDAD.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Viabilidad']

# Parámetros financieros
vp = ws['B8'].value        # 8,340,000 COP
tea = ws['C8'].value       # 13% (0.13)
i_mensual = ws['D8'].value # 1.0237% mensual (0.010237)
g = ws['E8'].value         # Gradiente 2% (0.02)
cuota_ini = ws['F8'].value # Cuota inicial 100,000 COP
periodo_yrs = ws['G8'].value # 10 años (120 meses)

print("=== PARÁMETROS FINANCIEROS DEL EXCEL ===")
print(f"Valor del Proyecto (VP): {vp:,} COP")
print(f"Tasa Efectiva Anual (TEA): {tea:.2%}")
print(f"Tasa Mensual Vencida (i): {i_mensual:.6%}")
print(f"Gradiente Geométrico Mensual (g): {g:.2%}")
print(f"Cuota Inicial: {cuota_ini:,} COP")
print(f"Periodo de Financiamiento: {periodo_yrs} años ({periodo_yrs*12} meses)")

# Analizar la fórmula de la Cuota 1 en el Excel:
# En la celda C15 del Excel, la fórmula es: =(B8-F8)/(G8*12)
# Es decir: (8,340,000 - 100,000) / 120 = 8,240,000 / 120 = 68,666.67 COP
# ¡ESTA FÓRMULA ES FINANCIERAMENTE INCORRECTA!
# Divide el saldo de la deuda (VP - Cuota Inicial) de forma lineal entre 120 meses para hallar la primera cuota,
# ignorando por completo el valor del dinero en el tiempo (intereses) y la fórmula del gradiente geométrico.
# 
# En un gradiente geométrico creciente, el valor presente de las cuotas debe ser igual a la deuda inicial (V_0 = VP - Cuota Inicial).
# La fórmula para el Valor Presente (VP) de un gradiente geométrico creciente es:
# Si g != i:
# VP_deuda = A * [ 1 - ((1 + g) / (1 + i))^n ] / (i - g)
# Donde:
# A = Primera cuota (Cuota 1)
# i = Tasa de interés mensual
# g = Gradiente geométrico mensual
# n = Número de periodos (120 meses)
#
# Calculemos cuál debería ser la Cuota 1 real (A) para que el crédito se liquide exactamente en la cuota 120.

n = periodo_yrs * 12
deuda_inicial = vp - cuota_ini

# Aplicando la fórmula de la primera cuota (A) para un gradiente geométrico creciente:
# deuda_inicial = A * [ 1 - ((1 + g) / (1 + i))^n ] / (i - g)
# A = deuda_inicial * (i - g) / [ 1 - ((1 + g) / (1 + i))^n ]

factor = (1 - ((1 + g) / (1 + i_mensual))**n) / (i_mensual - g)
cuota_1_real = deuda_inicial / factor

print("\n=== ANÁLISIS DE LA CUOTA 1 ===")
print(f"Cuota 1 calculada en Excel: {ws['C15'].value:,.2f} COP")
print(f"Cuota 1 REAL necesaria para amortizar la deuda: {cuota_1_real:,.2f} COP")

# Si usamos la cuota de Excel (68,666.67 COP), los intereses en el mes 1 son:
# Intereses = Deuda Inicial * i = 8,240,000 * 0.010237 = 84,352.88 COP (en el Excel dice 85,375.28 porque usa B8 completo en lugar de B8 - F8!)
# De hecho, en F14 del Excel dice =+B8 (8,340,000 COP) en lugar de B8 - F8 (8,240,000 COP)! O sea, ¡se les olvidó restar la cuota inicial de la deuda!
# Como la Cuota 1 (68,666.67) es menor que el interés del primer mes (84,352.88), la amortización es NEGATIVA:
# Amortización = Cuota - Interés = 68,666.67 - 84,352.88 = -15,686.21 COP.
# Esto hace que la deuda CREZCA en los primeros meses (capitalización de intereses o amortización negativa),
# lo cual es un comportamiento muy inusual y peligroso en créditos reales.
# Y debido a que el gradiente geométrico crece un 2% mensual, la cuota crece muy rápido.
# Veamos en qué cuota el saldo deudor del Excel se vuelve cero. El script anterior mostró que en la cuota 82 el saldo es negativo (-237,558.94).
# ¡Eso significa que el crédito se paga en 82 meses (6.8 años) en lugar de los 120 meses (10 años) planeados!
# ¿Por qué se paga antes? Porque la fórmula de la Cuota 1 en el Excel es un error garrafal, y al crecer un 2% mensual,
# las cuotas se vuelven gigantescas al final, amortizando la deuda mucho antes de lo previsto.
# Hagamos una tabla comparativa de lo que pasa en la cuota 80 en el Excel:
# Cuota 80 en Excel = 328,215.84 COP.
# Si el crédito durara los 120 meses con la fórmula de Excel, la cuota 120 sería:
# Cuota 120 = Cuota 1 * (1 + g)^119 = 68,666.67 * (1.02)^119 = 716,903.95 COP!
# Una cuota gigante para un pequeño productor.

# Calculemos el comportamiento con la amortización REAL del gradiente geométrico (cuota 1 = 44,705.54 COP)
# Con la cuota 1 real, la cuota 120 sería:
# Cuota 120 = 44,705.54 * (1.02)^119 = 466,734.50 COP.
# Y el saldo deudor llegaría exactamente a 0 en la cuota 120.

print("\n=== COMPARATIVA DE COMPORTAMIENTO ===")
print("Mes | Cuota Excel | Saldo Excel | Cuota Real Gradiente | Saldo Real Gradiente")
print("-" * 85)

saldo_excel = deuda_inicial + cuota_ini # El Excel arranca con 8,340,000 (no resta la cuota inicial!)
saldo_real = deuda_inicial # El real arranca con 8,240,000 (8,340,000 - 100,000)

cuota_excel = 68666.67
cuota_real = cuota_1_real

for mes in range(1, 121):
    # Excel
    if mes == 1:
        cuota_ex_mes = cuota_excel
    else:
        cuota_ex_mes = cuota_excel * (1 + g)**(mes - 1)
    
    int_ex_mes = saldo_excel * i_mensual
    amort_ex_mes = cuota_ex_mes - int_ex_mes
    saldo_excel_prev = saldo_excel
    saldo_excel -= amort_ex_mes
    
    # Real
    cuota_real_mes = cuota_real * (1 + g)**(mes - 1)
    int_real_mes = saldo_real * i_mensual
    amort_real_mes = cuota_real_mes - int_real_mes
    saldo_real -= amort_real_mes
    
    if mes in [1, 2, 10, 20, 40, 60, 80, 82, 100, 120]:
        se = f"{saldo_excel_prev - amort_ex_mes:,.2f}" if saldo_excel_prev > 0 else "PAGADO"
        sr = f"{saldo_real:,.2f}" if saldo_real > -1 else "0.00"
        print(f"{mes:3d} | {cuota_ex_mes:11,.2f} | {se:>11} | {cuota_real_mes:20,.2f} | {sr:>20}")
