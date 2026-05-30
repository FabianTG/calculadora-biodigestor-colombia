import openpyxl

path = "/home/ubuntu/upload/VIABILIDAD.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb['Viabilidad']

print("=== FÓRMULAS CLAVE ===")
# Celda del biogás estimado
print(f"C2 (Cabezas): {ws['C2'].value}")
print(f"D2 (Estiércol): {ws['D2'].value}")
print(f"E2 (Biogás): {ws['E2'].value}")
print(f"F2 (Viabilidad): {ws['F2'].value}")

print("\n=== PARÁMETROS FINANCIEROS ===")
print(f"C7 (Valor Proyecto): {ws['C7'].value}")
print(f"D7 (TEA): {ws['D7'].value}")
print(f"E7 (Tasa mensual vencida): {ws['E7'].value}")
print(f"F7 (Gradiente): {ws['F7'].value}")
print(f"G7 (Cuota inicial): {ws['G7'].value}")
print(f"H7 (Periodo): {ws['H7'].value}")

print("\n=== PRIMERAS FILAS DE LA TABLA ===")
for r in range(13, 20):
    row_vals = [ws.cell(row=r, column=c).value for column_idx, c in enumerate(range(2, 8))]
    print(f"Fila {r}: {row_vals}")
