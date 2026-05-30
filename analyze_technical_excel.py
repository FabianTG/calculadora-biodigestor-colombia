import openpyxl

path = "/home/ubuntu/upload/VIABILIDAD.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Viabilidad']

cabezas = ws['B3'].value
manure = ws['C3'].value
biogas = ws['D3'].value
viab = ws['E3'].value

print("=== PARÁMETROS TÉCNICOS EN EXCEL ===")
print(f"Cabezas de ganado: {cabezas}")
print(f"Estiércol diario (KG): {manure} (Fórmula: Cabezas * 8.9)")
print(f"Biogás estimado (M3/día): {biogas} (Fórmula: Estiércol * 0.17)")
print(f"Estado de viabilidad: {viab} (Fórmula: si cabezas >= 36 viable, >= 20 limitada, no viable)")

# Analicemos la física detrás de esto:
# 1. Estiércol por animal: 8.9 kg/animal/día.
# En el artículo científico, la producción fecal total de un bovino de doble propósito es de 40 kg/animal/día.
# El valor de 8.9 kg/animal/día corresponde a una fracción de recolección de aproximadamente 22.25% (40 * 0.2225 = 8.9 kg).
# ¡Es decir, asume una recolección fija muy baja (alrededor del 22%) sin justificarla dinámicamente!
#
# 2. Rendimiento de biogás: 0.17 m³/kg de estiércol.
# En el modelo real, el rendimiento de biogás se calcula como:
# Biogás = Estiércol (kg) * VS_FRACCION (0.12) * Rendimiento de Metano por piso térmico.
# Si el rendimiento de biogás por kg de estiércol fresco es 0.17 m³/kg, esto significaría:
# 0.17 = VS_FRACCION (0.12) * Rendimiento de Metano.
# Rendimiento de Metano = 0.17 / 0.12 = 1.416 m³/kg VS.
# ¡ESTO ES FÍSICAMENTE IMPOSIBLE! El rendimiento teórico máximo de metano del estiércol bovino es de 0.17 a 0.20 m³/kg VS (Sólidos Volátiles).
# El Excel multiplica directamente el estiércol fresco (kg) por 0.17 m³/kg para hallar el biogás, lo cual confunde:
# - Sólidos Volátiles (VS) con Estiércol Fresco.
# - Metano (CH4) con Biogás total.
# En la práctica, 1 kg de estiércol fresco produce alrededor de 0.015 a 0.020 m³ de biogás, NO 0.17 m³.
# Multiplicar por 0.17 directamente el estiércol fresco da una sobreestimación de casi 10 veces la producción real de biogás.
#
# Por ejemplo, para 36 vacas con 8.9 kg/vaca/día = 320.4 kg estiércol/día.
# En Excel da: 320.4 * 0.17 = 54.47 m³/día de biogás.
# En la realidad (piso térmico templado con factor 0.75 y rendimiento 0.1275 m³/kg VS):
# Biogás real = 320.4 * 0.12 * 0.1275 = 4.90 m³/día de biogás.
# ¡El Excel está calculando 54.47 m³/día, lo cual es 11 veces mayor de lo físicamente posible!
# Esto es un error crítico que haría que un productor construya un sistema gigante esperando un biogás que jamás se producirá.
